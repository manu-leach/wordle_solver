#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 22 18:04:52 2023

@author: cyleong
"""
import openpyxl
import wordle_clone
# Make a file named 'weighted-wordlist.txt' before using

def gen_weights(lex_name):

    lex_path = 'lexicons/{}.txt'.format(lex_name)
    
    lexicon = wordle_clone.Lexicon()
    lexicon.load_from_txt(lex_path)

    weight_dict = dict()
    for word in lexicon.word_list:
        weight_dict[word] = 1000.0

    dataframe = openpyxl.load_workbook("unigram_freq.xlsx")
    dataframe1 = dataframe.active   
    for row in range(1, dataframe1.max_row):
        word = str(dataframe1.cell(row= row, column= 1).value)         
        if word in lexicon.word_list:      
           weight_dict[word] = dataframe1.cell(row= row, column= 2).value
    
    value_sum = 0
    for value in weight_dict.values():
        value_sum += value

    for word, value in weight_dict.items():
        weight_dict[word] /= value_sum

    weighted_lex_path = 'lexicons/weighted_{}.txt'.format(lex_name)
    with open(weighted_lex_path, mode="wt", encoding="utf8") as ww:      
         for word, weight in weight_dict.items():
             ww.write("{} {}\n".format(word, weight))

def main():
    
    lexicon_name = 'sgb_words'
    gen_weights(lexicon_name)
    
if __name__ == '__main__':
    main()
          
          
          
          
          
          
          
    
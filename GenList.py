#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 22 18:04:52 2023

@author: cyleong
"""
import openpyxl
# Make a file named 'weighted-wordlist.txt' before using


def GenList():
    dict={}
    with open(("valid-wordle-words.txt"), mode="rt", encoding="utf8") as wlist: 
        for i in range(1,14855) :
          entry= wlist.readline().replace('\n','')
          dict[str(entry)] = "1000.0"

    dataframe = openpyxl.load_workbook("unigram_freq.xlsx")
    dataframe1 = dataframe.active   
    for row in range(1, dataframe1.max_row):
        word = str(dataframe1.cell(row= row, column= 1).value)
        lth=len(word)          
        if lth==5:      
           dict[ str(word) ] = str(dataframe1.cell(row= row, column= 2).value)
    
    with open(("weighted-wordlist.txt"), mode="wt", encoding="utf8") as ww:      
         for k in dict.keys():
             ww.write("{} {}\n".format(k, dict[k]))
         ww.close()
          
          
          
          
          
          
          
    